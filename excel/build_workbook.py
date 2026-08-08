import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

NAVY = "1F3864"
GOLD = "B08D57"
LIGHT = "EAF1F8"
WHITE = "FFFFFF"

rfm = pd.read_csv("../data/customer_rfm_churn.csv")

wb = Workbook()

# ---------------------------------------------------------------
# Sheet: Customer_RFM (source data)
# ---------------------------------------------------------------
ws_data = wb.active
ws_data.title = "Customer_RFM"
cols = ["CustomerID", "Country", "Recency", "Frequency", "Monetary", "Tenure",
        "AvgOrderValue", "CancellationRate", "Churned", "R_Score", "F_Score",
        "M_Score", "Segment", "ChurnProbability"]
ws_data.append(cols)
for c in ws_data[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for _, row in rfm.iterrows():
    ws_data.append([row[c] for c in cols])
for i in range(1, len(cols) + 1):
    ws_data.column_dimensions[get_column_letter(i)].width = 14
n_rows = len(rfm) + 1

# ---------------------------------------------------------------
# Sheet: Segment_Summary (formula-driven)
# ---------------------------------------------------------------
ws_seg = wb.create_sheet("Segment_Summary")
segments = ["Champions", "At-Risk High Value", "Regular", "Hibernating", "New / Low Engagement"]
ws_seg.append(["Segment", "Customers", "PctOfCustomers", "TotalRevenue", "AvgRecency"])
for c in ws_seg[1]:
    c.font = Font(bold=True, color=WHITE, name="Arial")
    c.fill = PatternFill("solid", fgColor=NAVY)
for i, seg in enumerate(segments):
    r = i + 2
    ws_seg.cell(row=r, column=1, value=seg)
    ws_seg.cell(row=r, column=2, value=f'=COUNTIF(Customer_RFM!$M$2:$M${n_rows},A{r})')
    ws_seg.cell(row=r, column=3, value=f'=ROUND(B{r}/{len(rfm)}*100,1)')
    ws_seg.cell(row=r, column=4, value=f'=ROUND(SUMIF(Customer_RFM!$M$2:$M${n_rows},A{r},Customer_RFM!$E$2:$E${n_rows}),0)')
    ws_seg.cell(row=r, column=5, value=f'=ROUND(AVERAGEIF(Customer_RFM!$M$2:$M${n_rows},A{r},Customer_RFM!$C$2:$C${n_rows}),0)')
for i in range(1, 6):
    ws_seg.column_dimensions[get_column_letter(i)].width = 20

# Country churn table
ws_seg.cell(row=1, column=7, value="Country").font = Font(bold=True, color=WHITE, name="Arial")
ws_seg.cell(row=1, column=8, value="Customers").font = Font(bold=True, color=WHITE, name="Arial")
ws_seg.cell(row=1, column=9, value="ChurnRatePct").font = Font(bold=True, color=WHITE, name="Arial")
for c in (7, 8, 9):
    ws_seg.cell(row=1, column=c).fill = PatternFill("solid", fgColor=GOLD)
countries = ["Portugal", "Ireland", "Netherlands", "United Kingdom", "France", "Spain", "Germany", "Belgium"]
for i, ctry in enumerate(countries):
    r = i + 2
    ws_seg.cell(row=r, column=7, value=ctry)
    ws_seg.cell(row=r, column=8, value=f'=COUNTIF(Customer_RFM!$B$2:$B${n_rows},G{r})')
    ws_seg.cell(row=r, column=9, value=f'=ROUND(SUMIFS(Customer_RFM!$I$2:$I${n_rows},Customer_RFM!$B$2:$B${n_rows},G{r})/H{r}*100,1)')
for i in (7, 8, 9):
    ws_seg.column_dimensions[get_column_letter(i)].width = 16

# ---------------------------------------------------------------
# Sheet: Dashboard
# ---------------------------------------------------------------
ws_d = wb.create_sheet("Dashboard", 0)
ws_d.sheet_view.showGridLines = False

ws_d.merge_cells("B2:K2")
ws_d["B2"] = "E-COMMERCE CUSTOMER CHURN DASHBOARD"
ws_d["B2"].font = Font(bold=True, size=20, color=NAVY, name="Arial")
ws_d.merge_cells("B3:K3")
ws_d["B3"] = "RFM Segmentation & Churn Risk — Dec 2010 to Dec 2011"
ws_d["B3"].font = Font(italic=True, size=12, color=GOLD, name="Arial")

def kpi_card(ws, col, label, formula, fmt="#,##0"):
    col_letter = get_column_letter(col)
    ws.merge_cells(f"{col_letter}5:{get_column_letter(col+1)}5")
    ws[f"{col_letter}5"] = label
    ws[f"{col_letter}5"].font = Font(bold=True, color=WHITE, size=11, name="Arial")
    ws[f"{col_letter}5"].fill = PatternFill("solid", fgColor=NAVY)
    ws[f"{col_letter}5"].alignment = Alignment(horizontal="center")
    ws.merge_cells(f"{col_letter}6:{get_column_letter(col+1)}7")
    cell = ws[f"{col_letter}6"]
    cell.value = formula
    cell.font = Font(bold=True, size=20, color=GOLD, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.number_format = fmt
    for rr in (5, 6, 7):
        for cc in (col, col + 1):
            ws.cell(row=rr, column=cc).border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

kpi_card(ws_d, 2, "TOTAL CUSTOMERS", f"=COUNTA(Customer_RFM!A2:A{n_rows})", "#,##0")
kpi_card(ws_d, 4, "OVERALL CHURN RATE", f"=ROUND(SUM(Customer_RFM!I2:I{n_rows})/COUNTA(Customer_RFM!A2:A{n_rows})*100,1)", '0.0"%"')
kpi_card(ws_d, 6, "AT-RISK HIGH VALUE", "=Segment_Summary!B3", "#,##0")
kpi_card(ws_d, 8, "REVENUE AT STAKE", "=Segment_Summary!D3", '#,##0" €"')
kpi_card(ws_d, 10, "IRELAND CHURN RATE", "=Segment_Summary!I3", '0.0"%"')

ws_d.row_dimensions[6].height = 22
ws_d.row_dimensions[7].height = 22

# Segment revenue bar chart
bar = BarChart()
bar.title = "Total Historical Revenue by RFM Segment"
bar.style = 10
bar.y_axis.title = "Revenue (€)"
data = Reference(ws_seg, min_col=4, min_row=1, max_row=6)
cats = Reference(ws_seg, min_col=1, min_row=2, max_row=6)
bar.add_data(data, titles_from_data=True)
bar.set_categories(cats)
bar.width, bar.height = 16, 9
ws_d.add_chart(bar, "B10")

# Country churn rate bar chart
bar2 = BarChart()
bar2.title = "Churn Rate by Country"
bar2.style = 11
bar2.y_axis.title = "Churn rate (%)"
data2 = Reference(ws_seg, min_col=9, min_row=1, max_row=9)
cats2 = Reference(ws_seg, min_col=7, min_row=2, max_row=9)
bar2.add_data(data2, titles_from_data=True)
bar2.set_categories(cats2)
bar2.width, bar2.height = 16, 9
ws_d.add_chart(bar2, "B29")

for i in range(1, 12):
    ws_d.column_dimensions[get_column_letter(i)].width = 16
ws_d.page_setup.orientation = "landscape"
ws_d.page_setup.fitToWidth = 1
ws_d.page_setup.fitToHeight = 0
ws_d.sheet_properties.pageSetUpPr.fitToPage = True

wb.save("Ecommerce_Churn_Dashboard.xlsx")
print("saved")
